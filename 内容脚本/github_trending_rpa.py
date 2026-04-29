#!/usr/bin/env python3
"""
GitHub Trending RPA 自动化脚本
使用 Playwright 录制完整流程，并抓取 README.md 内容存储到 content.db
"""

import re
from datetime import datetime
from pathlib import Path
import sys
import os

# 修复 Windows GBK 终端编码问题
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# 添加父目录到路径，以便导入项目模块
sys.path.insert(0, str(Path(__file__).parent.parent))

from config import CONTENT_DB_URL
from models.content import SessionLocal, Content

# 桌面路径
DESKTOP_PATH = Path(r"D:\桌面\新建文件夹\生产内容")
EXCEL_FILE = DESKTOP_PATH / f"{datetime.now().strftime('%Y-%m-%d')}_github_trending.xlsx"

def translate_to_chinese(texts: list[str]) -> list[str]:
    """批量翻译文本为中文，使用 Google Translate（免费，无需 Key）"""
    try:
        from deep_translator import GoogleTranslator
    except ImportError:
        print("[!] 未安装 deep-translator，简介将保留原文")
        print("    请安装: pip install deep-translator")
        return texts

    translator = GoogleTranslator(source="auto", target="zh-CN")
    results = []
    for text in texts:
        if not text.strip():
            results.append("")
            continue
        try:
            results.append(translator.translate(text))
        except Exception as e:
            print(f"  [!] 翻译失败，保留原文: {e}")
            results.append(text)
    return results


def fetch_readme_content(repo_name: str, page) -> str:
    """从 GitHub 仓库获取 README.md 内容"""
    try:
        readme_url = f"https://raw.githubusercontent.com/{repo_name}/main/README.md"
        response = page.context.request.get(readme_url)

        if response.status == 200:
            return response.text()

        # 尝试 master 分支
        readme_url = f"https://raw.githubusercontent.com/{repo_name}/master/README.md"
        response = page.context.request.get(readme_url)

        if response.status == 200:
            return response.text()

        return ""
    except Exception as e:
        print(f"    [!] 获取 README 失败 ({repo_name}): {e}")
        return ""


def save_to_content_db(repo_data: dict, readme_content: str) -> bool:
    """将抓取的数据保存到 content.db"""
    try:
        db = SessionLocal()

        # 检查是否已存在
        existing = db.query(Content).filter(
            Content.source_plat == "GitHub",
            Content.title == repo_data["项目名称"]
        ).first()

        if existing:
            # 更新现有记录
            existing.body = readme_content if readme_content else repo_data.get("简介（中文）", repo_data.get("简介", ""))
            existing.tags = f'["trending", "stars:{repo_data["累计Star"]}"]'
            db.commit()
            print(f"    [更新] {repo_data['项目名称']}")
        else:
            # 创建新记录
            content = Content(
                source_plat="GitHub",
                title=repo_data["项目名称"],
                body=readme_content if readme_content else repo_data.get("简介（中文）", repo_data.get("简介", "")),
                tags=f'["trending", "stars:{repo_data["累计Star"]}", "daily_increase:{repo_data["今日新增Star"]}"]'
            )
            db.add(content)
            db.commit()
            print(f"    [保存] {repo_data['项目名称']}")

        db.close()
        return True
    except Exception as e:
        print(f"    [错误] 保存到数据库失败: {e}")
        return False


def scrape_github_trending():
    """使用 Playwright 抓取 GitHub Trending 数据"""

    try:
        from playwright.sync_api import sync_playwright
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, Alignment
    except ImportError as e:
        print(f"[X] 缺少依赖: {e}")
        print("   请安装: pip install playwright openpyxl")
        print("   然后: playwright install chromium")
        return None

    trending_data = []

    with sync_playwright() as p:
        # 启动浏览器（无头模式，后台运行）
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        )
        page = context.new_page()

        try:
            # ========== Step 1: 访问 GitHub Trending ==========
            print("[Step 1] 访问 GitHub Trending...")
            page.goto("https://github.com/trending", wait_until="networkidle")
            page.wait_for_timeout(3000)

            # ========== Step 2: 等待页面加载 ==========
            print("[Step 2] 等待页面加载...")
            page.wait_for_selector("article.Box-row", timeout=10000)

            # ========== Step 3: 提取数据 ==========
            print("[Step 3] 提取 Trending 数据...")

            repo_elements = page.query_selector_all("article.Box-row")

            for i, repo in enumerate(repo_elements[:25], 1):
                try:
                    # 项目名称
                    name_elem = repo.query_selector("h2 a")
                    if not name_elem:
                        continue

                    full_name = name_elem.get_attribute("href").strip("/")

                    # 项目简介
                    desc_elem = repo.query_selector("p.col-9")
                    description = ""
                    if desc_elem:
                        description = desc_elem.inner_text().strip()

                    # 今日新增 Star
                    today_stars_elem = repo.query_selector("span.d-inline-block.float-sm-right")
                    today_stars = ""
                    if today_stars_elem:
                        today_stars_text = today_stars_elem.inner_text().strip()
                        match = re.search(r'([\d,]+)', today_stars_text)
                        if match:
                            today_stars = match.group(1).replace(",", "")

                    # 累计 Star 总数
                    stars_elem = repo.query_selector('a[href$="/stargazers"]')
                    total_stars = ""
                    if stars_elem:
                        stars_text = stars_elem.inner_text().strip()
                        match = re.search(r'([\d,\.]+)([kKmM]?)', stars_text)
                        if match:
                            num = match.group(1).replace(",", "")
                            suffix = match.group(2).lower()
                            if suffix == 'k':
                                total_stars = str(int(float(num) * 1000))
                            elif suffix == 'm':
                                total_stars = str(int(float(num) * 1000000))
                            else:
                                total_stars = num

                    repo_item = {
                        "项目名称": full_name,
                        "简介": description,
                        "今日新增Star": today_stars,
                        "累计Star": total_stars
                    }

                    # ========== Step 4: 抓取 README.md 内容 ==========
                    print(f"  [{i}] {full_name} - 今日+{today_stars}⭐ 累计{total_stars}⭐")
                    print(f"      [获取README中...]")

                    readme_content = fetch_readme_content(full_name, page)

                    if readme_content:
                        print(f"      [✓] README 已获取 ({len(readme_content)} 字符)")
                    else:
                        print(f"      [·] README 为空，将使用简介")

                    # ========== Step 5: 保存到数据库 ==========
                    save_to_content_db(repo_item, readme_content)

                    trending_data.append(repo_item)

                except Exception as e:
                    print(f"  [!] 解析项目 {i} 时出错: {e}")
                    continue

            browser.close()

        except Exception as e:
            print(f"[X] 抓取过程中出错: {e}")
            browser.close()
            return None

    return trending_data


def write_to_excel(data):
    """将数据写入 Excel 文件"""

    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        print("[X] 请先安装 openpyxl: pip install openpyxl")
        return False

    today = datetime.now().strftime("%Y-%m-%d")

    # 检查文件是否存在
    if EXCEL_FILE.exists():
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "GitHub Trending"

        # 写入表头
        headers = ["项目名称", "简介（原文）", "简介（中文）", "今日新增Star", "累计Star"]
        ws.append(headers)

        # 设置表头样式
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    # 添加日期分隔行
    ws.append([f"=== {today} ===", "", "", ""])

    # 写入数据
    for item in data:
        ws.append([
            item["项目名称"],
            item["简介"],
            item["简介（中文）"],
            item["今日新增Star"],
            item["累计Star"]
        ])

    # 调整列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    # 保存文件
    wb.save(EXCEL_FILE)
    print(f"\n[OK] 数据已保存到: {EXCEL_FILE}")
    print(f"   共写入 {len(data)} 个项目")

    return True


def main():
    """主函数"""
    print("=" * 60)
    print("GitHub Trending RPA 自动化脚本")
    print("=" * 60)

    # 抓取数据
    data = scrape_github_trending()

    if not data:
        print("\n[X] 抓取失败，没有获取到数据")
        return 1

    print(f"\n[OK] 成功抓取 {len(data)} 个项目")

    # 翻译简介
    print("\n[Step 5] 翻译项目简介...")
    descriptions = [item["简介"] for item in data]
    translations = translate_to_chinese(descriptions)
    for item, zh in zip(data, translations):
        item["简介（中文）"] = zh

    # 写入 Excel
    if write_to_excel(data):
        return 0
    else:
        return 1


if __name__ == "__main__":
    exit(main())
